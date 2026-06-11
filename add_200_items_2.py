from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

# ===================== Gmail compose URL builder =====================
def build_gmail(email, maker_name, product_name):
    subject = "Potential Distribution Partnership for Japan"
    body = f"""Dear {maker_name} Team,

My name is Hiroyuki Inoguchi from
Sumai pluS Co., Ltd. in Japan

I hope this message finds you all well.

Our company focuses on promoting products that enrich people's daily lives. We are currently seeking unique international brands to introduce to the Japanese market and support their growth.

I recently had the opportunity to review your product ({product_name}) and was very impressed by its potential in the Japanese market.

I am confident that your product will strongly appeal to Japanese customers, and I would like to explore the possibility of building a partnership with your company to help ensure its success.

Furthermore, our company has a proven track record of promoting overseas products through our proprietary sales network and Japanese distribution channels.

Would it be possible to schedule a brief online meeting sometime next week to discuss this matter?

I look forward to hearing from you.


Sincerely,

Hiroyuki Inoguchi
Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )

================================================

Sumai pluS Co., Ltd.
 (LEAGUE Co., Ltd. Agent )
Hiroyuki Inoguchi

E-mail:yutorin.ino@gmail.com

Address: 49-5 Kitazakuno, Higashi-Itsushiro, Ichinomiya City, Aichi Prefecture, Japan

================================================="""
    return f"https://mail.google.com/mail/?view=cm&fs=1&to={quote(email,safe='')}&su={quote(subject,safe='')}&body={quote(body,safe='')}"

# ===================== 200 new products (No.906-1105) =====================
new_products = [
    # ── キッチン・調理器具 ──────────────────────────────
    {"no":906,"category":"キッチン・調理器具","name":"FryAway Plant-Based Cooking Oil Solidifier","maker":"FryAway","ec_site":"FryAway公式サイト","prod_url":"https://fryaway.co/collections/all","maker_hp":"https://fryaway.co","contact":"info@fryaway.co"},
    {"no":907,"category":"キッチン・調理器具","name":"Brisk It Origin 580 AI Smart Pellet Grill","maker":"Brisk It","ec_site":"Brisk It公式サイト","prod_url":"https://briskitgrills.com/products/origin-580","maker_hp":"https://briskitgrills.com","contact":"support@briskitgrills.com"},
    {"no":908,"category":"キッチン・調理器具","name":"Agari Smart Cooker Precision Oven","maker":"Agari Kitchen","ec_site":"Agari Kitchen公式サイト","prod_url":"https://agarikitchen.com/products/agari-smart-cooker","maker_hp":"https://agarikitchen.com","contact":"info@agarikitchen.com"},
    {"no":909,"category":"キッチン・調理器具","name":"Tovala Smart Oven Pro","maker":"Tovala","ec_site":"Tovala公式サイト","prod_url":"https://tovala.com/tovala-smart-oven/","maker_hp":"https://tovala.com","contact":"hello@tovala.com"},
    {"no":910,"category":"キッチン・調理器具","name":"Combustion Inc Predictive Thermometer","maker":"Combustion Inc","ec_site":"Combustion Inc公式サイト","prod_url":"https://combustion.inc/products/predictive-thermometer","maker_hp":"https://combustion.inc","contact":"hello@combustion.inc"},
    {"no":911,"category":"キッチン・調理器具","name":"MeatStick 4X Wireless Meat Thermometer","maker":"The MeatStick","ec_site":"MeatStick公式サイト","prod_url":"https://themeatstick.com/collections/all","maker_hp":"https://themeatstick.com","contact":"hello@themeatstick.com"},
    {"no":912,"category":"キッチン・調理器具","name":"Solo Stove Pi Pizza Oven","maker":"Solo Stove","ec_site":"Solo Stove公式サイト","prod_url":"https://www.solostove.com/en-us/c/pizza-ovens/","maker_hp":"https://www.solostove.com","contact":"support@solostove.com"},
    {"no":913,"category":"キッチン・調理器具","name":"Brava Pure Light Oven","maker":"Brava","ec_site":"Brava公式サイト","prod_url":"https://www.brava.com/pages/the-oven","maker_hp":"https://www.brava.com","contact":"support@brava.com"},
    {"no":914,"category":"キッチン・調理器具","name":"Anova Precision Chamber Vacuum Sealer Pro","maker":"Anova Culinary","ec_site":"Anova公式サイト","prod_url":"https://anovaculinary.com/products/anova-precision-vacuum-sealer-pro","maker_hp":"https://anovaculinary.com","contact":"support@anovaculinary.com"},
    {"no":915,"category":"キッチン・調理器具","name":"Fellow Stagg EKG+ Electric Pour-Over Kettle","maker":"Fellow Products","ec_site":"Fellow公式サイト","prod_url":"https://fellowproducts.com/products/staggekg","maker_hp":"https://fellowproducts.com","contact":"support@fellowproducts.com"},
    {"no":916,"category":"キッチン・調理器具","name":"Stasher Reusable Silicone Storage Bag Set","maker":"Stasher","ec_site":"Stasher公式サイト","prod_url":"https://www.stasherbag.com/collections/all-bags","maker_hp":"https://www.stasherbag.com","contact":"hello@stasherbag.com"},
    {"no":917,"category":"キッチン・調理器具","name":"Lékué Microwave Steam Case Set","maker":"Lékué","ec_site":"Lékué公式サイト","prod_url":"https://www.lekue.com/en/steam-cases/","maker_hp":"https://www.lekue.com","contact":"info@lekue.com"},
    {"no":918,"category":"キッチン・調理器具","name":"Dreamfarm Supoon Sit Up Scraping Spoon","maker":"Dreamfarm","ec_site":"Dreamfarm公式サイト","prod_url":"https://dreamfarm.com/products/supoon","maker_hp":"https://dreamfarm.com","contact":"hello@dreamfarm.com"},
    {"no":919,"category":"キッチン・調理器具","name":"Full Circle Breeze Odor-Free Compost Collector","maker":"Full Circle","ec_site":"Full Circle公式サイト","prod_url":"https://www.fullcirclelife.com/collections/compost","maker_hp":"https://www.fullcirclelife.com","contact":"info@fullcirclelife.com"},
    {"no":920,"category":"キッチン・調理器具","name":"Kuhn Rikon Pull Chop Food Chopper","maker":"Kuhn Rikon","ec_site":"Kuhn Rikon公式サイト","prod_url":"https://www.kuhnrikon.com/us/products/choppers/","maker_hp":"https://www.kuhnrikon.com","contact":"info@kuhnrikon.com"},
    {"no":921,"category":"キッチン・調理器具","name":"Zyliss Easy Pull Manual Food Processor","maker":"Zyliss","ec_site":"Zyliss公式サイト","prod_url":"https://www.zyliss.com/collections/food-processors","maker_hp":"https://www.zyliss.com","contact":"info@zyliss.com"},
    {"no":922,"category":"キッチン・調理器具","name":"Zoku Classic Pop Maker","maker":"Zoku","ec_site":"Zoku公式サイト","prod_url":"https://www.zokuhome.com/collections/ice-pop-makers","maker_hp":"https://www.zokuhome.com","contact":"info@zokuhome.com"},
    {"no":923,"category":"キッチン・調理器具","name":"Joseph Joseph Chop2Pot Folding Chopping Board","maker":"Joseph Joseph","ec_site":"Joseph Joseph公式サイト","prod_url":"https://www.josephjoseph.com/collections/chopping-boards","maker_hp":"https://www.josephjoseph.com","contact":"help@josephjoseph.com"},
    {"no":924,"category":"キッチン・調理器具","name":"Prepd Pack Smart Lunchbox System","maker":"Prepd","ec_site":"Prepd公式サイト","prod_url":"https://www.prepd.in/collections/all","maker_hp":"https://www.prepd.in","contact":"info@prepd.in"},
    {"no":925,"category":"キッチン・調理器具","name":"GreenPan Reserve Ceramic Nonstick Cookware Set","maker":"GreenPan","ec_site":"GreenPan公式サイト","prod_url":"https://www.greenpan.us/collections/reserve","maker_hp":"https://www.greenpan.us","contact":"info@greenpanusa.com"},
    # ── スマートホーム・インテリア・照明 ───────────────────
    {"no":926,"category":"スマートホーム・インテリア・照明","name":"BedJet 3 Dual Zone Climate Comfort System","maker":"BedJet","ec_site":"BedJet公式サイト","prod_url":"https://bedjet.com/products/bedjet-3-dual-zone-climate-comfort-system-for-couples","maker_hp":"https://bedjet.com","contact":"info@bedjet.com"},
    {"no":927,"category":"スマートホーム・インテリア・照明","name":"Flume 2 Smart Home Water Monitor","maker":"Flume Water","ec_site":"Flume公式サイト","prod_url":"https://flumewater.com/product/","maker_hp":"https://flumewater.com","contact":"support@flumewater.com"},
    {"no":928,"category":"スマートホーム・インテリア・照明","name":"Airthings View Plus Air Quality Monitor","maker":"Airthings","ec_site":"Airthings公式サイト","prod_url":"https://www.airthings.com/view-plus","maker_hp":"https://www.airthings.com","contact":"support@airthings.com"},
    {"no":929,"category":"スマートホーム・インテリア・照明","name":"Graywind Motorized Smart Roller Shades","maker":"Graywind","ec_site":"Graywind公式サイト","prod_url":"https://www.graywind.com/collections/smart-shades","maker_hp":"https://www.graywind.com","contact":"sales@graywind.com"},
    {"no":930,"category":"スマートホーム・インテリア・照明","name":"Nanoleaf Skylight Ceiling Light Panels","maker":"Nanoleaf","ec_site":"Nanoleaf公式サイト","prod_url":"https://nanoleaf.me/en-US/products/smarter-kits/","maker_hp":"https://nanoleaf.me","contact":"support@nanoleaf.me"},
    {"no":931,"category":"スマートホーム・インテリア・照明","name":"Level Lock+ Connect Smart Lock","maker":"Level Home","ec_site":"Level Home公式サイト","prod_url":"https://level.co/products/lock-plus","maker_hp":"https://level.co","contact":"hello@level.co"},
    {"no":932,"category":"スマートホーム・インテリア・照明","name":"Flair Smart Vent System","maker":"Flair Smart Home","ec_site":"Flair公式サイト","prod_url":"https://flair.co/collections/products","maker_hp":"https://flair.co","contact":"hello@flair.co"},
    {"no":933,"category":"スマートホーム・インテリア・照明","name":"Emporia Vue 3 Home Energy Monitor","maker":"Emporia Energy","ec_site":"Emporia公式サイト","prod_url":"https://www.emporiaenergy.com/vue-3","maker_hp":"https://www.emporiaenergy.com","contact":"hello@emporiaenergy.com"},
    {"no":934,"category":"スマートホーム・インテリア・照明","name":"LIFX Beam Smart LED Light Bar","maker":"LIFX","ec_site":"LIFX公式サイト","prod_url":"https://www.lifx.com/collections/beam","maker_hp":"https://www.lifx.com","contact":"support@lifx.com"},
    {"no":935,"category":"スマートホーム・インテリア・照明","name":"Meross Smart Garage Door Opener","maker":"Meross","ec_site":"Meross公式サイト","prod_url":"https://www.meross.com/en-us/smart-home/garage-door-openers/","maker_hp":"https://www.meross.com","contact":"support@meross.com"},
    {"no":936,"category":"スマートホーム・インテリア・照明","name":"Netatmo Smart Carbon Monoxide Alarm","maker":"Netatmo","ec_site":"Netatmo公式サイト","prod_url":"https://www.netatmo.com/en-us/security/co-alarm","maker_hp":"https://www.netatmo.com","contact":"support@netatmo.com"},
    {"no":937,"category":"スマートホーム・インテリア・照明","name":"YoLink Smart Home Sensor Starter Kit","maker":"YoSmart","ec_site":"YoLink公式サイト","prod_url":"https://www.yosmart.com/collections/yolink","maker_hp":"https://www.yosmart.com","contact":"support@yosmart.com"},
    {"no":938,"category":"スマートホーム・インテリア・照明","name":"Govee Glide Hexa Pro Smart LED Panels","maker":"Govee","ec_site":"Govee公式サイト","prod_url":"https://us.govee.com/collections/led-panels","maker_hp":"https://us.govee.com","contact":"support@govee.com"},
    {"no":939,"category":"スマートホーム・インテリア・照明","name":"Tempest Smart Home Weather Station","maker":"WeatherFlow Tempest","ec_site":"Tempest公式サイト","prod_url":"https://shop.tempest.earth/products/tempest","maker_hp":"https://tempest.earth","contact":"https://tempest.earth/contact/（問い合わせページ）"},
    {"no":940,"category":"スマートホーム・インテリア・照明","name":"Ecowitt GW2001 Professional Weather Station","maker":"Ecowitt","ec_site":"Ecowitt公式サイト","prod_url":"https://www.ecowitt.com/shop/product/list/id/14.html","maker_hp":"https://www.ecowitt.com","contact":"support@ecowitt.com"},
    {"no":941,"category":"スマートホーム・インテリア・照明","name":"Rachio 3 Smart Sprinkler Controller 16-Zone","maker":"Rachio","ec_site":"Rachio公式サイト","prod_url":"https://rachio.com/products/rachio-3","maker_hp":"https://rachio.com","contact":"https://rachio.com/pages/support（問い合わせページ）"},
    {"no":942,"category":"スマートホーム・インテリア・照明","name":"DeepMag Levitating Moon Lamp","maker":"DeepMag","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=DeepMag+levitating+lamp","maker_hp":"https://www.deepmag.com","contact":"info@deepmag.com"},
    {"no":943,"category":"スマートホーム・インテリア・照明","name":"SANDSARA Kinetic Sand Art Coffee Table","maker":"SANDSARA","ec_site":"SANDSARA公式サイト","prod_url":"https://www.sandsara.io/products","maker_hp":"https://www.sandsara.io","contact":"info@sandsara.io"},
    {"no":944,"category":"スマートホーム・インテリア・照明","name":"Poplight Tool-Free Renter-Friendly Wall Light","maker":"Poplight","ec_site":"Poplight公式サイト","prod_url":"https://www.poplightco.com","maker_hp":"https://www.poplightco.com","contact":"hello@poplightco.com"},
    {"no":945,"category":"スマートホーム・インテリア・照明","name":"Objet Charger 35W Home-Décor USB-C Charger","maker":"Objet Devices","ec_site":"Objet公式サイト","prod_url":"https://objetdevices.com","maker_hp":"https://objetdevices.com","contact":"hello@objetdevices.com"},
    # ── ウェアラブル・ヘルス・フィットネス ──────────────────
    {"no":946,"category":"ウェアラブル・ヘルス・フィットネス","name":"Ultrahuman Ring AIR Smart Ring","maker":"Ultrahuman","ec_site":"Ultrahuman公式サイト","prod_url":"https://www.ultrahuman.com/ring/","maker_hp":"https://www.ultrahuman.com","contact":"support@ultrahuman.com"},
    {"no":947,"category":"ウェアラブル・ヘルス・フィットネス","name":"RingConn Gen 2 Smart Health Ring","maker":"RingConn","ec_site":"RingConn公式サイト","prod_url":"https://ringconn.com/products/ringconn-gen-2","maker_hp":"https://ringconn.com","contact":"cs@ringconn.com"},
    {"no":948,"category":"ウェアラブル・ヘルス・フィットネス","name":"WHOOP 4.0 Performance Health Tracker","maker":"WHOOP","ec_site":"WHOOP公式サイト","prod_url":"https://www.whoop.com/us/en/","maker_hp":"https://www.whoop.com","contact":"support@whoop.com"},
    {"no":949,"category":"ウェアラブル・ヘルス・フィットネス","name":"Oura Ring 4 Smart Ring","maker":"Oura Ring","ec_site":"Oura公式サイト","prod_url":"https://ouraring.com/product/rings/heritage","maker_hp":"https://ouraring.com","contact":"support@ouraring.com"},
    {"no":950,"category":"ウェアラブル・ヘルス・フィットネス","name":"COROS NOMAD GPS Watch","maker":"COROS","ec_site":"COROS公式サイト","prod_url":"https://us.coros.com/products/coros-nomad","maker_hp":"https://us.coros.com","contact":"support@coros.com"},
    {"no":951,"category":"ウェアラブル・ヘルス・フィットネス","name":"Elemind Sleep Headband Neurotech Wearable","maker":"Elemind Technologies","ec_site":"Elemind公式サイト","prod_url":"https://elemindtech.com/products/elemind","maker_hp":"https://elemindtech.com","contact":"support@elemindtech.com"},
    {"no":952,"category":"ウェアラブル・ヘルス・フィットネス","name":"Kokoon Nightbuds Sleep Earbuds","maker":"Kokoon","ec_site":"Kokoon公式サイト","prod_url":"https://kokoon.io/en-us/products/nightbuds-sleep-headphones","maker_hp":"https://kokoon.io","contact":"help@kokoon.io"},
    {"no":953,"category":"ウェアラブル・ヘルス・フィットネス","name":"Withings ScanWatch 2 Hybrid Smartwatch","maker":"Withings","ec_site":"Withings公式サイト","prod_url":"https://www.withings.com/us/en/scanwatch","maker_hp":"https://www.withings.com","contact":"support@withings.com"},
    {"no":954,"category":"ウェアラブル・ヘルス・フィットネス","name":"Withings Sleep Analyzer Under-Mattress Sensor","maker":"Withings","ec_site":"Withings公式サイト","prod_url":"https://www.withings.com/us/en/sleep","maker_hp":"https://www.withings.com","contact":"support@withings.com"},
    {"no":955,"category":"ウェアラブル・ヘルス・フィットネス","name":"DUSQ Sleep Regulation Behind-Ear Wearable","maker":"DUSQ","ec_site":"DUSQ公式サイト","prod_url":"https://dusq.com","maker_hp":"https://dusq.com","contact":"info@dusq.com"},
    {"no":956,"category":"ウェアラブル・ヘルス・フィットネス","name":"Circular Ring 2 Smart Health Tracker","maker":"Circular","ec_site":"Circular公式サイト","prod_url":"https://www.circular.xyz/shop","maker_hp":"https://www.circular.xyz","contact":"support@circular.xyz"},
    {"no":957,"category":"ウェアラブル・ヘルス・フィットネス","name":"Amazfit Active 2 Smartwatch","maker":"Amazfit","ec_site":"Amazfit公式サイト","prod_url":"https://www.amazfit.com/en/active-2.html","maker_hp":"https://www.amazfit.com","contact":"support@amazfit.com"},
    {"no":958,"category":"ウェアラブル・ヘルス・フィットネス","name":"Garmin Enduro 3 Ultraperformance GPS Watch","maker":"Garmin","ec_site":"Garmin公式サイト","prod_url":"https://www.garmin.com/en-US/p/enduro-3","maker_hp":"https://www.garmin.com","contact":"support@garmin.com"},
    {"no":959,"category":"ウェアラブル・ヘルス・フィットネス","name":"Plaud Note Pro AI Voice Recorder","maker":"Plaud.ai","ec_site":"Plaud公式サイト","prod_url":"https://www.plaud.ai/products/plaud-note-pro","maker_hp":"https://www.plaud.ai","contact":"support@plaud.ai"},
    {"no":960,"category":"ウェアラブル・ヘルス・フィットネス","name":"Oxiline Pulse XS Pro Blood Oxygen Monitor","maker":"Oxiline","ec_site":"Oxiline公式サイト","prod_url":"https://www.oxiline.com/products/pulse-xs-pro","maker_hp":"https://www.oxiline.com","contact":"support@oxiline.com"},
    # ── アウトドア・スポーツ・旅行 ────────────────────────
    {"no":961,"category":"アウトドア・スポーツ・旅行","name":"PACKFIRE Foldable Portable Fire Pit","maker":"PACKFIRE","ec_site":"PACKFIRE公式サイト","prod_url":"https://packfire.com/products/packfire","maker_hp":"https://packfire.com","contact":"info@packfire.com"},
    {"no":962,"category":"アウトドア・スポーツ・旅行","name":"Graphene-X Aerograph Puffer Jacket","maker":"Graphene-X","ec_site":"Graphene-X公式サイト","prod_url":"https://graphene-x.com/collections/jackets","maker_hp":"https://graphene-x.com","contact":"support@graphene-x.com"},
    {"no":963,"category":"アウトドア・スポーツ・旅行","name":"SENIQ Trail System Hiking Apparel","maker":"SENIQ","ec_site":"SENIQ公式サイト","prod_url":"https://seniqbrand.com/collections/hike","maker_hp":"https://seniqbrand.com","contact":"hello@seniqbrand.com"},
    {"no":964,"category":"アウトドア・スポーツ・旅行","name":"Pingora Ultralight Backpack 38L","maker":"Pingora","ec_site":"Pingora公式サイト","prod_url":"https://pingora-gear.com/collections/hiking-backpacks","maker_hp":"https://pingora-gear.com","contact":"https://pingora-gear.com/pages/contact（問い合わせページ）"},
    {"no":965,"category":"アウトドア・スポーツ・旅行","name":"BougeRV 200W Portable Solar Panel Foldable","maker":"BougeRV","ec_site":"BougeRV公式サイト","prod_url":"https://www.bougerv.com/collections/portable-solar-panels","maker_hp":"https://www.bougerv.com","contact":"hello@bougerv.com"},
    {"no":966,"category":"アウトドア・スポーツ・旅行","name":"Garmin InReach Messenger Plus Satellite Communicator","maker":"Garmin","ec_site":"Garmin公式サイト","prod_url":"https://www.garmin.com/en-US/p/inreach-messenger-plus","maker_hp":"https://www.garmin.com","contact":"support@garmin.com"},
    {"no":967,"category":"アウトドア・スポーツ・旅行","name":"NEMO Tensor Elite Ultralight Sleeping Pad","maker":"NEMO Equipment","ec_site":"NEMO公式サイト","prod_url":"https://www.nemoequipment.com/collections/sleeping-pads","maker_hp":"https://www.nemoequipment.com","contact":"info@nemoequipment.com"},
    {"no":968,"category":"アウトドア・スポーツ・旅行","name":"Outdoor Element Flicker Feather Fixed-Blade Knife","maker":"Outdoor Element","ec_site":"Outdoor Element公式サイト","prod_url":"https://outdoorelement.com/collections/knives","maker_hp":"https://outdoorelement.com","contact":"info@outdoorelement.com"},
    {"no":969,"category":"アウトドア・スポーツ・旅行","name":"Sea to Summit Ultralight XR Sleeping Mat","maker":"Sea to Summit","ec_site":"Sea to Summit公式サイト","prod_url":"https://seatosummit.com/collections/sleeping-mats","maker_hp":"https://seatosummit.com","contact":"customerservice@seatosummit.com"},
    {"no":970,"category":"アウトドア・スポーツ・旅行","name":"Katadyn BeFree AC Activated Carbon Water Filter Bottle","maker":"Katadyn","ec_site":"Katadyn公式サイト","prod_url":"https://www.katadyn.com/en-us/collections/filters","maker_hp":"https://www.katadyn.com","contact":"info@katadyn.com"},
    {"no":971,"category":"アウトドア・スポーツ・旅行","name":"EcoFlow Wave 3 Portable Air Conditioner","maker":"EcoFlow","ec_site":"EcoFlow公式サイト","prod_url":"https://www.ecoflow.com/products/ecoflow-wave-3","maker_hp":"https://www.ecoflow.com","contact":"service@ecoflow.com"},
    {"no":972,"category":"アウトドア・スポーツ・旅行","name":"Jackery Explorer 1000 Plus Power Station","maker":"Jackery","ec_site":"Jackery公式サイト","prod_url":"https://www.jackery.com/products/explorer-1000-plus","maker_hp":"https://www.jackery.com","contact":"support@jackery.com"},
    {"no":973,"category":"アウトドア・スポーツ・旅行","name":"Thule ReVert Hitch-Mount Vertical Bike Rack","maker":"Thule","ec_site":"Thule公式サイト","prod_url":"https://www.thule.com/en-us/bike-rack","maker_hp":"https://www.thule.com","contact":"support.us@thule.com"},
    {"no":974,"category":"アウトドア・スポーツ・旅行","name":"Coast EAL35R Voice-Activated Camp Lantern","maker":"Coast","ec_site":"Coast公式サイト","prod_url":"https://www.coastportland.com/collections/lanterns","maker_hp":"https://www.coastportland.com","contact":"info@coastportland.com"},
    {"no":975,"category":"アウトドア・スポーツ・旅行","name":"Grubcan Carbon Kevlar Bear Canister 6.6L","maker":"Grubcan","ec_site":"Grubcan公式サイト","prod_url":"https://grubcan.com","maker_hp":"https://grubcan.com","contact":"hello@grubcan.com"},
    # ── ペット用品 ─────────────────────────────────────
    {"no":976,"category":"ペット用品","name":"Enabot ROLA PetTracker GPS Camera Collar","maker":"Enabot","ec_site":"Enabot公式サイト","prod_url":"https://www.enabot.com/products/rola-pettracker-gps-tracker-for-pets","maker_hp":"https://www.enabot.com","contact":"info@enabot.com"},
    {"no":977,"category":"ペット用品","name":"Fi Series 3 Smart GPS Dog Collar","maker":"Fi","ec_site":"Fi公式サイト","prod_url":"https://tryfi.com/collections/collars","maker_hp":"https://tryfi.com","contact":"support@tryfi.com"},
    {"no":978,"category":"ペット用品","name":"Tractive GPS & Health Tracker for Dogs","maker":"Tractive","ec_site":"Tractive公式サイト","prod_url":"https://tractive.com/en/shop/dog-gps-trackers","maker_hp":"https://tractive.com","contact":"https://help.tractive.com/hc/en-us/requests/new（問い合わせページ）"},
    {"no":979,"category":"ペット用品","name":"PupPod Rocker Dog Enrichment Game System","maker":"PupPod","ec_site":"PupPod公式サイト","prod_url":"https://puppod.com/products/puppod-rocker-with-feeder","maker_hp":"https://puppod.com","contact":"info@puppod.com"},
    {"no":980,"category":"ペット用品","name":"Petcube Bites 2 Lite Pet Camera & Treat Dispenser","maker":"Petcube","ec_site":"Petcube公式サイト","prod_url":"https://petcube.com/bites2lite.html","maker_hp":"https://petcube.com","contact":"support@petcube.com"},
    {"no":981,"category":"ペット用品","name":"Furbo 360° Rotating Dog Camera","maker":"Furbo","ec_site":"Furbo公式サイト","prod_url":"https://www.furbo.com/products/furbo-360-dog-camera","maker_hp":"https://www.furbo.com","contact":"support@furbo.com"},
    {"no":982,"category":"ペット用品","name":"Whistle Switch Smart Dog Health Tracker","maker":"Whistle Labs","ec_site":"Whistle公式サイト","prod_url":"https://www.whistle.com/products/","maker_hp":"https://www.whistle.com","contact":"support@whistle.com"},
    {"no":983,"category":"ペット用品","name":"Wagz Explore Smart Dog Collar","maker":"Wagz","ec_site":"Wagz公式サイト","prod_url":"https://wagz.com/products/collar","maker_hp":"https://wagz.com","contact":"hello@wagz.com"},
    {"no":984,"category":"ペット用品","name":"GROOMI Vacuum Deshedding Dog Brush","maker":"The GROOMI","ec_site":"GROOMI公式サイト","prod_url":"https://www.thegroomi.com/en-us/products/groomi","maker_hp":"https://www.thegroomi.com","contact":"hello@thegroomi.com"},
    {"no":985,"category":"ペット用品","name":"Petlibro Automatic Cat Feeder with Camera","maker":"Petlibro","ec_site":"Petlibro公式サイト","prod_url":"https://petlibro.com/collections/automatic-feeders","maker_hp":"https://petlibro.com","contact":"support@petlibro.com"},
    {"no":986,"category":"ペット用品","name":"Dogness Smart IPCam Dog Feeder","maker":"Dogness","ec_site":"Dogness公式サイト","prod_url":"https://www.dogness.com/collections/feeders","maker_hp":"https://www.dogness.com","contact":"service@dogness.com"},
    {"no":987,"category":"ペット用品","name":"Air Doctor 4-in-1 HEPA Pet Air Purifier","maker":"Air Doctor","ec_site":"Air Doctor公式サイト","prod_url":"https://www.airdoctorpro.com/products/ad5000","maker_hp":"https://www.airdoctorpro.com","contact":"hello@airdoctorpro.com"},
    {"no":988,"category":"ペット用品","name":"Litter-Robot 4 Automatic Self-Cleaning Litter Box","maker":"Litter-Robot","ec_site":"Litter-Robot公式サイト","prod_url":"https://www.litter-robot.com/litter-robot-4","maker_hp":"https://www.litter-robot.com","contact":"support@litter-robot.com"},
    {"no":989,"category":"ペット用品","name":"Sure Petcare Microchip Cat Door Connect","maker":"Sure Petcare","ec_site":"Sure Petcare公式サイト","prod_url":"https://www.surepetcare.com/en-us/cat-doors","maker_hp":"https://www.surepetcare.com","contact":"support@surepetcare.com"},
    {"no":990,"category":"ペット用品","name":"SpotOn Virtual Fence GPS Dog Collar","maker":"SpotOn","ec_site":"SpotOn公式サイト","prod_url":"https://spotoncollar.com/products/","maker_hp":"https://spotoncollar.com","contact":"support@spotoncollar.com"},
    # ── テクノロジー・ガジェット ────────────────────────
    {"no":991,"category":"テクノロジー・ガジェット","name":"HOVERAir X1 PROMAX 8K Self-Flying Camera Drone","maker":"HOVERAir","ec_site":"HOVERAir公式サイト","prod_url":"https://hoverair.com/pages/x1-pro-and-promax","maker_hp":"https://hoverair.com","contact":"support@hoverair.com"},
    {"no":992,"category":"テクノロジー・ガジェット","name":"Plaud Note Pro AI Voice Recorder Card","maker":"Plaud.ai","ec_site":"Plaud公式サイト","prod_url":"https://www.plaud.ai/products/plaud-note-pro","maker_hp":"https://www.plaud.ai","contact":"support@plaud.ai"},
    {"no":993,"category":"テクノロジー・ガジェット","name":"XREAL Air 2 Pro AR Smart Glasses","maker":"XREAL","ec_site":"XREAL公式サイト","prod_url":"https://www.xreal.com/air2pro/","maker_hp":"https://www.xreal.com","contact":"support@xreal.com"},
    {"no":994,"category":"テクノロジー・ガジェット","name":"Rabbit R1 Pocket AI Companion Device","maker":"Rabbit Inc","ec_site":"Rabbit公式サイト","prod_url":"https://www.rabbit.tech/rabbit-r1","maker_hp":"https://www.rabbit.tech","contact":"hello@rabbit.tech"},
    {"no":995,"category":"テクノロジー・ガジェット","name":"Rewind Pendant Wearable AI Memory Device","maker":"Rewind AI","ec_site":"Rewind公式サイト","prod_url":"https://www.rewind.ai/pendant","maker_hp":"https://www.rewind.ai","contact":"support@rewind.ai"},
    {"no":996,"category":"テクノロジー・ガジェット","name":"Brilliant Smart Home Control Panel","maker":"Brilliant","ec_site":"Brilliant公式サイト","prod_url":"https://www.brilliant.tech/products/","maker_hp":"https://www.brilliant.tech","contact":"hello@brilliant.tech"},
    {"no":997,"category":"テクノロジー・ガジェット","name":"iSee Intelligent TV Camera for Video Calls","maker":"iSee Tech","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=iSee+TV+camera","maker_hp":"https://iseetech.com","contact":"support@iseetech.com"},
    {"no":998,"category":"テクノロジー・ガジェット","name":"Ridge Magnetic Power Bank 10000mAh","maker":"Ridge Wallet","ec_site":"Ridge公式サイト","prod_url":"https://ridge.com/collections/power-banks","maker_hp":"https://ridge.com","contact":"support@ridge.com"},
    {"no":999,"category":"テクノロジー・ガジェット","name":"Aulumu G05 Pro 5-in-1 Magnetic Phone Stand","maker":"Aulumu","ec_site":"Aulumu公式サイト","prod_url":"https://aulumu.com/products/g05-pro","maker_hp":"https://aulumu.com","contact":"hello@aulumu.com"},
    {"no":1000,"category":"テクノロジー・ガジェット","name":"Anker SOLIX F3800 Portable Power Station","maker":"Anker","ec_site":"Anker公式サイト","prod_url":"https://www.anker.com/products/solix-f3800-portable-power-station","maker_hp":"https://www.anker.com","contact":"support@anker.com"},
    {"no":1001,"category":"テクノロジー・ガジェット","name":"Logitech MX Creative Console Customizable Keypad","maker":"Logitech","ec_site":"Logitech公式サイト","prod_url":"https://www.logitech.com/en-us/products/keyboards/mx-creative-console.html","maker_hp":"https://www.logitech.com","contact":"support@logitech.com"},
    {"no":1002,"category":"テクノロジー・ガジェット","name":"BOOX Go Air 4 Color E-Ink Tablet","maker":"ONYX BOOX","ec_site":"BOOX公式サイト","prod_url":"https://shop.boox.com/products/goair4c","maker_hp":"https://www.boox.com","contact":"support@boox.com"},
    {"no":1003,"category":"テクノロジー・ガジェット","name":"Kensington StudioDock iPad Docking Station","maker":"Kensington","ec_site":"Kensington公式サイト","prod_url":"https://www.kensington.com/p/products/electronic-accessories/ipad-docking-stations/studiodock/","maker_hp":"https://www.kensington.com","contact":"support@kensington.com"},
    {"no":1004,"category":"テクノロジー・ガジェット","name":"Nanit Pro Baby Monitor with Breathing Wear","maker":"Nanit","ec_site":"Nanit公式サイト","prod_url":"https://www.nanit.com/products/nanit-pro","maker_hp":"https://www.nanit.com","contact":"support@nanit.com"},
    {"no":1005,"category":"テクノロジー・ガジェット","name":"LockPickingLawyer-Edition Titanium Multi-Tool","maker":"Prometheus Design Werx","ec_site":"PDW公式サイト","prod_url":"https://prometheusdesignwerx.com/collections/multi-tools","maker_hp":"https://prometheusdesignwerx.com","contact":"info@prometheusdesignwerx.com"},
    # ── 美容・スキンケア ───────────────────────────────
    {"no":1006,"category":"美容・スキンケア","name":"FLOSSOM FR100 Pro RF & Microcurrent Beauty Device","maker":"FLOSSOM","ec_site":"FLOSSOM公式サイト","prod_url":"https://flossom.com/products/lift-firm-rf-beauty-device","maker_hp":"https://flossom.com","contact":"hello@flossom.com"},
    {"no":1007,"category":"美容・スキンケア","name":"Currentbody Skin LED Light Therapy Mask","maker":"CurrentBody","ec_site":"CurrentBody公式サイト","prod_url":"https://www.currentbody.com/products/currentbody-skin-led-light-therapy-face-mask","maker_hp":"https://www.currentbody.com","contact":"info@currentbody.com"},
    {"no":1008,"category":"美容・スキンケア","name":"Nira Precision Anti-Aging Laser","maker":"Nira Skincare","ec_site":"Nira公式サイト","prod_url":"https://niraskincare.com/products/nira-precision-anti-aging-laser","maker_hp":"https://niraskincare.com","contact":"support@niraskincare.com"},
    {"no":1009,"category":"美容・スキンケア","name":"Nuface Trinity+ Facial Toning Device","maker":"NuFACE","ec_site":"NuFACE公式サイト","prod_url":"https://www.mynuface.com/products/nuface-trinity-plus","maker_hp":"https://www.mynuface.com","contact":"support@mynuface.com"},
    {"no":1010,"category":"美容・スキンケア","name":"Solawave 4-in-1 Red Light Wand","maker":"Solawave","ec_site":"Solawave公式サイト","prod_url":"https://solawave.co/products/radiant-renewal-skincare-wand","maker_hp":"https://solawave.co","contact":"hello@solawave.co"},
    {"no":1011,"category":"美容・スキンケア","name":"Foreo Bear 2 Microcurrent Facial Device","maker":"Foreo","ec_site":"Foreo公式サイト","prod_url":"https://www.foreo.com/products/bear-2","maker_hp":"https://www.foreo.com","contact":"support@foreo.com"},
    {"no":1012,"category":"美容・スキンケア","name":"Therabody TheraFace Pro Facial Device","maker":"Therabody","ec_site":"Therabody公式サイト","prod_url":"https://www.therabody.com/us/en-us/theraface-pro.html","maker_hp":"https://www.therabody.com","contact":"support@therabody.com"},
    {"no":1013,"category":"美容・スキンケア","name":"Glo Skin Beauty Pro Microderm Exfoliator","maker":"Glo Skin Beauty","ec_site":"Glo Skin Beauty公式サイト","prod_url":"https://www.gloskinbeauty.com/collections/skincare-tools","maker_hp":"https://www.gloskinbeauty.com","contact":"info@gloskinbeauty.com"},
    {"no":1014,"category":"美容・スキンケア","name":"Project E Beauty Ultrasonic Skin Scrubber","maker":"Project E Beauty","ec_site":"Project E Beauty公式サイト","prod_url":"https://projectebeauty.com/collections/facial-tools","maker_hp":"https://projectebeauty.com","contact":"support@projectebeauty.com"},
    {"no":1015,"category":"美容・スキンケア","name":"Amiro R3 Turbo RF Skin Tightening Device","maker":"Amiro","ec_site":"Amiro公式サイト","prod_url":"https://amiro.com/collections/rf-beauty-devices","maker_hp":"https://amiro.com","contact":"support@amiro.com"},
    {"no":1016,"category":"美容・スキンケア","name":"Medicube Age-R Ultra Tune EMS Face Massager","maker":"Medicube","ec_site":"Medicube公式サイト","prod_url":"https://www.medicube.us/products/age-r-ultra-tune","maker_hp":"https://www.medicube.us","contact":"help@medicube.us"},
    {"no":1017,"category":"美容・スキンケア","name":"Silk'n ReVit Microdermabrasion Device","maker":"Silk'n","ec_site":"Silk'n公式サイト","prod_url":"https://www.silkn.com/collections/microdermabrasion","maker_hp":"https://www.silkn.com","contact":"support@silkn.com"},
    {"no":1018,"category":"美容・スキンケア","name":"Trophy Skin BrightenMD Microdermabrasion System","maker":"Trophy Skin","ec_site":"Trophy Skin公式サイト","prod_url":"https://www.trophyskin.com/products/brightenmd","maker_hp":"https://www.trophyskin.com","contact":"support@trophyskin.com"},
    {"no":1019,"category":"美容・スキンケア","name":"Ziip Nano Current Facial Device","maker":"ZIIP Beauty","ec_site":"ZIIP公式サイト","prod_url":"https://ziipbeauty.com/products/halo","maker_hp":"https://ziipbeauty.com","contact":"support@ziipbeauty.com"},
    {"no":1020,"category":"美容・スキンケア","name":"Trufacial Home Facial Rejuvenation Device","maker":"Trufacial","ec_site":"Trufacial公式サイト","prod_url":"https://trufacial.com/products","maker_hp":"https://trufacial.com","contact":"hello@trufacial.com"},
    # ── 子供・教育 ─────────────────────────────────────
    {"no":1021,"category":"子供・教育","name":"Haivivi BubblePal AI Interactive Toy Companion","maker":"Haivivi","ec_site":"Haivivi公式サイト","prod_url":"https://www.haivivi.com/en/product","maker_hp":"https://www.haivivi.com","contact":"customerservice@haivivi.com"},
    {"no":1022,"category":"子供・教育","name":"iYofe 24V Kids Electric Ride-On Car","maker":"iYofe","ec_site":"iYofe公式サイト","prod_url":"https://iyofe.com/collections/all","maker_hp":"https://iyofe.com","contact":"info@iyofe.com"},
    {"no":1023,"category":"子供・教育","name":"FoloToy AI Conversational Plush Bear","maker":"FoloToy","ec_site":"FoloToy公式サイト","prod_url":"https://store.folotoy.com/collections/all","maker_hp":"https://folotoy.com","contact":"https://store.folotoy.com/pages/contact（問い合わせページ）"},
    {"no":1024,"category":"子供・教育","name":"Toniebox Audio Player for Children Stories","maker":"Boxine","ec_site":"Toniebox公式サイト","prod_url":"https://us.tonies.com/products/toniebox","maker_hp":"https://us.tonies.com","contact":"support@tonies.com"},
    {"no":1025,"category":"子供・教育","name":"Osmo Genius Starter Kit iPad Learning System","maker":"Osmo","ec_site":"Osmo公式サイト","prod_url":"https://www.playosmo.com/en/starter-kit-ipad/","maker_hp":"https://www.playosmo.com","contact":"support@playosmo.com"},
    {"no":1026,"category":"子供・教育","name":"Kano Harry Potter Coding Kit","maker":"Kano","ec_site":"Kano公式サイト","prod_url":"https://kano.me/store/en-us/collections/kits","maker_hp":"https://kano.me","contact":"support@kano.me"},
    {"no":1027,"category":"子供・教育","name":"Sphero BOLT App-Enabled Robot Ball","maker":"Sphero","ec_site":"Sphero公式サイト","prod_url":"https://sphero.com/products/sphero-bolt","maker_hp":"https://sphero.com","contact":"support@sphero.com"},
    {"no":1028,"category":"子供・教育","name":"Coding express train set robot","maker":"STEAM Town","ec_site":"STEAM Town公式サイト","prod_url":"https://www.steamtown.com","maker_hp":"https://www.steamtown.com","contact":"info@steamtown.com"},
    {"no":1029,"category":"子供・教育","name":"myFirst Fone R2s Kids GPS Smartwatch","maker":"myFirst","ec_site":"myFirst公式サイト","prod_url":"https://www.myfirst.com/products/myfirst-fone-r2s","maker_hp":"https://www.myfirst.com","contact":"support@myfirst.com"},
    {"no":1030,"category":"子供・教育","name":"Turing Tumble Mechanical Computer Board Game","maker":"Turing Tumble","ec_site":"Turing Tumble公式サイト","prod_url":"https://turingtumble.com/products/turing-tumble","maker_hp":"https://turingtumble.com","contact":"info@turingtumble.com"},
    {"no":1031,"category":"子供・教育","name":"Cosmo Robot by Anki (re-launch) Interactive AI Robot","maker":"Digital Dream Labs","ec_site":"Digital Dream Labs公式サイト","prod_url":"https://www.digitaldreamlabs.com/collections/cozmo","maker_hp":"https://www.digitaldreamlabs.com","contact":"hello@digitaldreamlabs.com"},
    {"no":1032,"category":"子供・教育","name":"Artie Max Robot That Teaches Kids to Code","maker":"Educational Insights","ec_site":"Educational Insights公式サイト","prod_url":"https://www.educationalinsights.com/artie-max","maker_hp":"https://www.educationalinsights.com","contact":"service@educationalinsights.com"},
    {"no":1033,"category":"子供・教育","name":"Learning Resources Coding Critters Ranger","maker":"Learning Resources","ec_site":"Learning Resources公式サイト","prod_url":"https://www.learningresources.com/coding-critters","maker_hp":"https://www.learningresources.com","contact":"info@learningresources.com"},
    {"no":1034,"category":"子供・教育","name":"Wonder Workshop Dash Robot Coding Toy","maker":"Wonder Workshop","ec_site":"Wonder Workshop公式サイト","prod_url":"https://www.makewonder.com/robots/dash","maker_hp":"https://www.makewonder.com","contact":"support@makewonder.com"},
    {"no":1035,"category":"子供・教育","name":"littleBits Rule Your Room Smart Home Kit","maker":"littleBits","ec_site":"littleBits公式サイト","prod_url":"https://littlebits.com/collections/kits","maker_hp":"https://littlebits.com","contact":"info@littlebits.com"},
    # ── ファッション・アクセサリー ──────────────────────
    {"no":1036,"category":"ファッション・アクセサリー","name":"Cotopaxi Batac 24L Backpack Del Dia","maker":"Cotopaxi","ec_site":"Cotopaxi公式サイト","prod_url":"https://www.cotopaxi.com/collections/backpacks","maker_hp":"https://www.cotopaxi.com","contact":"help@cotopaxi.com"},
    {"no":1037,"category":"ファッション・アクセサリー","name":"Ridge Wallet Titanium Minimalist Card Holder","maker":"Ridge","ec_site":"Ridge公式サイト","prod_url":"https://ridge.com/collections/wallets","maker_hp":"https://ridge.com","contact":"support@ridge.com"},
    {"no":1038,"category":"ファッション・アクセサリー","name":"Dango D01 Pioneer Bifold Wallet with Multi-Tool","maker":"Dango Products","ec_site":"Dango公式サイト","prod_url":"https://dangoproducts.com/collections/wallets","maker_hp":"https://dangoproducts.com","contact":"info@dangoproducts.com"},
    {"no":1039,"category":"ファッション・アクセサリー","name":"Nomad Base One Max MagSafe Wireless Charger","maker":"Nomad Goods","ec_site":"Nomad公式サイト","prod_url":"https://nomadgoods.com/collections/charging","maker_hp":"https://nomadgoods.com","contact":"support@nomadgoods.com"},
    {"no":1040,"category":"ファッション・アクセサリー","name":"Peak Design Everyday Backpack 30L V2","maker":"Peak Design","ec_site":"Peak Design公式サイト","prod_url":"https://www.peakdesign.com/collections/backpacks","maker_hp":"https://www.peakdesign.com","contact":"help@peakdesign.com"},
    {"no":1041,"category":"ファッション・アクセサリー","name":"Moment iPhone MagSafe Strap Photo Case","maker":"Moment","ec_site":"Moment公式サイト","prod_url":"https://www.shopmoment.com/collections/cases","maker_hp":"https://www.shopmoment.com","contact":"support@shopmoment.com"},
    {"no":1042,"category":"ファッション・アクセサリー","name":"Fossil Gen 6 Hybrid HR Smartwatch","maker":"Fossil","ec_site":"Fossil公式サイト","prod_url":"https://www.fossil.com/en-us/wearables/","maker_hp":"https://www.fossil.com","contact":"fcs@fossil.com"},
    {"no":1043,"category":"ファッション・アクセサリー","name":"Bellroy Slim Sleeve Wallet","maker":"Bellroy","ec_site":"Bellroy公式サイト","prod_url":"https://bellroy.com/products/slim-sleeve-wallet","maker_hp":"https://bellroy.com","contact":"help@bellroy.com"},
    {"no":1044,"category":"ファッション・アクセサリー","name":"Aer City Pack Pro 2 Urban Backpack 24.6L","maker":"Aer","ec_site":"Aer公式サイト","prod_url":"https://www.aersf.com/collections/backpacks","maker_hp":"https://www.aersf.com","contact":"hello@aersf.com"},
    {"no":1045,"category":"ファッション・アクセサリー","name":"Saddleback Leather Thin Bifold Wallet","maker":"Saddleback Leather","ec_site":"Saddleback公式サイト","prod_url":"https://saddlebackleather.com/collections/wallets","maker_hp":"https://saddlebackleather.com","contact":"support@saddlebackleather.com"},
    {"no":1046,"category":"ファッション・アクセサリー","name":"Ekster Senate Slim Cardholder Wallet","maker":"Ekster","ec_site":"Ekster公式サイト","prod_url":"https://ekster.com/collections/wallets","maker_hp":"https://ekster.com","contact":"support@ekster.com"},
    {"no":1047,"category":"ファッション・アクセサリー","name":"Herschel Little America Backpack 25L","maker":"Herschel Supply Co","ec_site":"Herschel公式サイト","prod_url":"https://www.herschel.com/collections/backpacks","maker_hp":"https://www.herschel.com","contact":"customerservice@herschel.com"},
    {"no":1048,"category":"ファッション・アクセサリー","name":"Trayvax Element Stainless Steel Wallet","maker":"Trayvax","ec_site":"Trayvax公式サイト","prod_url":"https://trayvax.com/collections/wallets","maker_hp":"https://trayvax.com","contact":"support@trayvax.com"},
    {"no":1049,"category":"ファッション・アクセサリー","name":"Twelve South HiRise 3 Deluxe iPhone Stand","maker":"Twelve South","ec_site":"Twelve South公式サイト","prod_url":"https://www.twelvesouth.com/products/hirise-3-deluxe","maker_hp":"https://www.twelvesouth.com","contact":"support@twelvesouth.com"},
    {"no":1050,"category":"ファッション・アクセサリー","name":"Fjallraven Kanken Mini 7L Backpack","maker":"Fjallraven","ec_site":"Fjallraven公式サイト","prod_url":"https://www.fjallraven.com/us/en-us/bags-gear/kanken","maker_hp":"https://www.fjallraven.com","contact":"customer.service@fjallraven.com"},
    # ── クリーニング・収納・整理 ───────────────────────
    {"no":1051,"category":"クリーニング・収納・整理","name":"Hedgehog Dryer GO Portable Gear Dryer","maker":"Hedgehog Dryer","ec_site":"Hedgehog Dryer公式サイト","prod_url":"https://us.hedgehogdryer.com/products/hedgehog-go-new","maker_hp":"https://us.hedgehogdryer.com","contact":"contact@hedgehogdryer.com"},
    {"no":1052,"category":"クリーニング・収納・整理","name":"Townew T1S Smart Self-Sealing Trash Can","maker":"Townew","ec_site":"Townew公式サイト","prod_url":"https://thetownew.com/collections/all","maker_hp":"https://thetownew.com","contact":"oversea@townew.com"},
    {"no":1053,"category":"クリーニング・収納・整理","name":"Bissell CrossWave HF3 Cordless Wet Dry Vacuum","maker":"Bissell","ec_site":"Bissell公式サイト","prod_url":"https://www.bissell.com/products/crosswave","maker_hp":"https://www.bissell.com","contact":"consumercare@bissell.com"},
    {"no":1054,"category":"クリーニング・収納・整理","name":"Steamfast SF-375 Steam Mop with Pads","maker":"Steamfast","ec_site":"Steamfast公式サイト","prod_url":"https://steamfast.com/collections/steam-mops","maker_hp":"https://steamfast.com","contact":"support@steamfast.com"},
    {"no":1055,"category":"クリーニング・収納・整理","name":"DropStop Car Seat Gap Filler Organizer","maker":"DropStop","ec_site":"DropStop公式サイト","prod_url":"https://www.dropstopco.com/products/","maker_hp":"https://www.dropstopco.com","contact":"info@dropstopco.com"},
    {"no":1056,"category":"クリーニング・収納・整理","name":"Zulay Kitchen Over-Sink Dish Drying Rack","maker":"Zulay Kitchen","ec_site":"Zulay Kitchen公式サイト","prod_url":"https://www.zulaykitchen.com/collections/dish-racks","maker_hp":"https://www.zulaykitchen.com","contact":"support@zulaykitchen.com"},
    {"no":1057,"category":"クリーニング・収納・整理","name":"Umbra Trigg Floating Shelf & Display Ledge","maker":"Umbra","ec_site":"Umbra公式サイト","prod_url":"https://www.umbra.com/collections/shelves","maker_hp":"https://www.umbra.com","contact":"customerservice@umbra.com"},
    {"no":1058,"category":"クリーニング・収納・整理","name":"mDesign Stackable Drawer Organizer Set","maker":"mDesign","ec_site":"mDesign公式サイト","prod_url":"https://mdesign.com/collections/drawer-organizers","maker_hp":"https://mdesign.com","contact":"info@mdesign.com"},
    {"no":1059,"category":"クリーニング・収納・整理","name":"Rubbermaid Brilliance Leak-Proof Container Set","maker":"Rubbermaid","ec_site":"Rubbermaid公式サイト","prod_url":"https://www.rubbermaid.com/collections/food-storage","maker_hp":"https://www.rubbermaid.com","contact":"consumer@newell.com"},
    {"no":1060,"category":"クリーニング・収納・整理","name":"Simplehuman Sensor Pump XL 9oz Liquid Dispenser","maker":"simplehuman","ec_site":"simplehuman公式サイト","prod_url":"https://www.simplehuman.com/collections/sensor-pumps","maker_hp":"https://www.simplehuman.com","contact":"support@simplehuman.com"},
    {"no":1061,"category":"クリーニング・収納・整理","name":"Drawer Organizer Velvet Tray Set by Acrylic","maker":"DECOMOMO","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=DECOMOMO+velvet+organizer","maker_hp":"https://www.decomomo.com","contact":"info@decomomo.com"},
    {"no":1062,"category":"クリーニング・収納・整理","name":"OXO Good Grips Pop Container Set 8-Piece","maker":"OXO","ec_site":"OXO公式サイト","prod_url":"https://www.oxo.com/categories/organization","maker_hp":"https://www.oxo.com","contact":"info@oxo.com"},
    {"no":1063,"category":"クリーニング・収納・整理","name":"Casdon Dyson Cordless Vacuum Toy","maker":"Casdon","ec_site":"Casdon公式サイト","prod_url":"https://www.casdon.com/collections/vacuums","maker_hp":"https://www.casdon.com","contact":"info@casdon.com"},
    {"no":1064,"category":"クリーニング・収納・整理","name":"Smartake 20-Pack Velvet Hangers Non-Slip","maker":"SMARTAKE","ec_site":"Amazon","prod_url":"https://www.amazon.com/s?k=SMARTAKE+velvet+hangers","maker_hp":"https://smartake.com","contact":"support@smartake.com"},
    {"no":1065,"category":"クリーニング・収納・整理","name":"EasyBreathe Anti-Pollen Window Screen Filter","maker":"EasyBreathe","ec_site":"EasyBreathe公式サイト","prod_url":"https://easybreathe.com","maker_hp":"https://easybreathe.com","contact":"info@easybreathe.com"},
    # ── キッチン続き ──────────────────────────────────
    {"no":1066,"category":"キッチン・調理器具","name":"Ember Travel Mug 2+ Smart Temperature Mug","maker":"Ember","ec_site":"Ember公式サイト","prod_url":"https://ember.com/products/ember-travel-mug-2-plus","maker_hp":"https://ember.com","contact":"support@ember.com"},
    {"no":1067,"category":"キッチン・調理器具","name":"Modernist Cuisine Piezano Crispy Crust Pizza Oven","maker":"Piezano","ec_site":"Piezano公式サイト","prod_url":"https://piezanopizzaoven.com","maker_hp":"https://piezanopizzaoven.com","contact":"info@piezanopizzaoven.com"},
    {"no":1068,"category":"キッチン・調理器具","name":"Ooni Fyra 12 Wood-Fired Portable Pizza Oven","maker":"Ooni","ec_site":"Ooni公式サイト","prod_url":"https://ooni.com/products/ooni-fyra-12","maker_hp":"https://ooni.com","contact":"support@ooni.com"},
    {"no":1069,"category":"キッチン・調理器具","name":"Lomi Smart Countertop Composter","maker":"Pela Earth","ec_site":"Lomi公式サイト","prod_url":"https://lomi.com/products/lomi","maker_hp":"https://lomi.com","contact":"support@pela.earth"},
    {"no":1070,"category":"キッチン・調理器具","name":"Clever Coffee Dripper Complete Bundle","maker":"Abid Coffee Products","ec_site":"Abid公式サイト","prod_url":"https://cleverdrip.com/products","maker_hp":"https://cleverdrip.com","contact":"info@cleverdrip.com"},
    # ── スマートホーム続き ─────────────────────────────
    {"no":1071,"category":"スマートホーム・インテリア・照明","name":"Lexi Light Cordless Lamp with Rechargeable Battery","maker":"Lexi Lighting","ec_site":"Lexi公式サイト","prod_url":"https://www.lexilighting.com","maker_hp":"https://www.lexilighting.com","contact":"info@lexilighting.com"},
    {"no":1072,"category":"スマートホーム・インテリア・照明","name":"SwitchBot Hub 2 Smart IR Remote Controller","maker":"SwitchBot","ec_site":"SwitchBot公式サイト","prod_url":"https://www.switch-bot.com/products/switchbot-hub-2","maker_hp":"https://www.switch-bot.com","contact":"support@switch-bot.com"},
    {"no":1073,"category":"スマートホーム・インテリア・照明","name":"MOEN Smart Touchless Kitchen Faucet","maker":"Moen","ec_site":"Moen公式サイト","prod_url":"https://www.moen.com/smart-faucets","maker_hp":"https://www.moen.com","contact":"consumer.service@moen.com"},
    {"no":1074,"category":"スマートホーム・インテリア・照明","name":"Ultraloq U-Bolt Pro WiFi Smart Lock","maker":"Ultraloq","ec_site":"Ultraloq公式サイト","prod_url":"https://www.u-tec.com/collections/smart-locks","maker_hp":"https://www.u-tec.com","contact":"support@u-tec.com"},
    {"no":1075,"category":"スマートホーム・インテリア・照明","name":"August Wi-Fi Smart Lock 4th Gen","maker":"August Smart Lock","ec_site":"August公式サイト","prod_url":"https://august.com/products/august-wifi-smart-lock","maker_hp":"https://august.com","contact":"support@august.com"},
    # ── ウェアラブル続き ───────────────────────────────
    {"no":1076,"category":"ウェアラブル・ヘルス・フィットネス","name":"Fossil Gen 6 Wellness Edition Smartwatch","maker":"Fossil","ec_site":"Fossil公式サイト","prod_url":"https://www.fossil.com/en-us/wearables/gen-6/","maker_hp":"https://www.fossil.com","contact":"fcs@fossil.com"},
    {"no":1077,"category":"ウェアラブル・ヘルス・フィットネス","name":"Amazfit Balance Smartwatch 14-Day Battery","maker":"Amazfit","ec_site":"Amazfit公式サイト","prod_url":"https://www.amazfit.com/en/balance","maker_hp":"https://www.amazfit.com","contact":"support@amazfit.com"},
    {"no":1078,"category":"ウェアラブル・ヘルス・フィットネス","name":"Pebble Index 01 Smart Ring with Microphone","maker":"Pebble","ec_site":"Pebble公式サイト","prod_url":"https://www.pebble.com","maker_hp":"https://www.pebble.com","contact":"hello@pebble.com"},
    {"no":1079,"category":"ウェアラブル・ヘルス・フィットネス","name":"Renpho LYNX Smart Health Ring","maker":"Renpho","ec_site":"Renpho公式サイト","prod_url":"https://renpho.com/collections/smart-ring","maker_hp":"https://renpho.com","contact":"support@renpho.com"},
    {"no":1080,"category":"ウェアラブル・ヘルス・フィットネス","name":"Jabra Enhance Plus OTC Hearing Aids","maker":"Jabra","ec_site":"Jabra公式サイト","prod_url":"https://www.jabra.com/hearing","maker_hp":"https://www.jabra.com","contact":"support@jabra.com"},
    # ── アウトドア続き ─────────────────────────────────
    {"no":1081,"category":"アウトドア・スポーツ・旅行","name":"Biolite CampStove 2+ Wood-Burning Camping Stove","maker":"BioLite Energy","ec_site":"BioLite公式サイト","prod_url":"https://www.bioliteenergy.com/products/campstove-2-plus","maker_hp":"https://www.bioliteenergy.com","contact":"support@bioliteenergy.com"},
    {"no":1082,"category":"アウトドア・スポーツ・旅行","name":"MPOWERD Luci Outdoor 2.0 Solar Lantern","maker":"MPOWERD","ec_site":"MPOWERD公式サイト","prod_url":"https://mpowerd.com/collections/lanterns","maker_hp":"https://mpowerd.com","contact":"hello@mpowerd.com"},
    {"no":1083,"category":"アウトドア・スポーツ・旅行","name":"Grayl GeoPress Water Purifier Bottle","maker":"GRAYL","ec_site":"GRAYL公式サイト","prod_url":"https://grayl.com/products/geopress-purifier","maker_hp":"https://grayl.com","contact":"info@grayl.com"},
    {"no":1084,"category":"アウトドア・スポーツ・旅行","name":"Oru Beach LT Folding Kayak","maker":"Oru Kayak","ec_site":"Oru Kayak公式サイト","prod_url":"https://www.orukayak.com/collections/kayaks","maker_hp":"https://www.orukayak.com","contact":"hello@orukayak.com"},
    {"no":1085,"category":"アウトドア・スポーツ・旅行","name":"Lifestraw Go Series Water Filter Bottle 22oz","maker":"LifeStraw","ec_site":"LifeStraw公式サイト","prod_url":"https://lifestraw.com/products/lifestraw-go","maker_hp":"https://lifestraw.com","contact":"support@lifestraw.com"},
    # ── ペット続き ─────────────────────────────────────
    {"no":1086,"category":"ペット用品","name":"Waggle Pet Monitor Temperature & Humidity","maker":"Waggle","ec_site":"Waggle公式サイト","prod_url":"https://waggle.dog/products/waggle-pro","maker_hp":"https://waggle.dog","contact":"support@waggle.dog"},
    {"no":1087,"category":"ペット用品","name":"Halo 3+ GPS Wireless Dog Fence Collar","maker":"Halo Collar","ec_site":"Halo Collar公式サイト","prod_url":"https://www.halocollar.com/products/halo-collar-3","maker_hp":"https://www.halocollar.com","contact":"support@halocollar.com"},
    {"no":1088,"category":"ペット用品","name":"Clever Pet Hub Dog Computer Game Console","maker":"CleverPet","ec_site":"CleverPet公式サイト","prod_url":"https://cleverpet.com","maker_hp":"https://cleverpet.com","contact":"hello@cleverpet.com"},
    {"no":1089,"category":"ペット用品","name":"Moonwalkers Dog Treadmill PetRun PR720F","maker":"GoPet","ec_site":"GoPet公式サイト","prod_url":"https://gopetpetproducts.com/collections/dog-treadmill","maker_hp":"https://gopetpetproducts.com","contact":"support@gopetpetproducts.com"},
    {"no":1090,"category":"ペット用品","name":"Animo Activity and Behavior Monitor for Dogs","maker":"Sure Petcare","ec_site":"Sure Petcare公式サイト","prod_url":"https://www.surepetcare.com/en-us/animo","maker_hp":"https://www.surepetcare.com","contact":"support@surepetcare.com"},
    # ── テクノロジー続き ──────────────────────────────
    {"no":1091,"category":"テクノロジー・ガジェット","name":"Snapmaker Artisan 3-in-1 3D Printer CNC Laser","maker":"Snapmaker","ec_site":"Snapmaker公式サイト","prod_url":"https://snapmaker.com/snapmaker-artisan","maker_hp":"https://snapmaker.com","contact":"support@snapmaker.com"},
    {"no":1092,"category":"テクノロジー・ガジェット","name":"Bambu Lab P1S 3D Printer Fully Enclosed","maker":"Bambu Lab","ec_site":"Bambu Lab公式サイト","prod_url":"https://bambulab.com/en-us/p1","maker_hp":"https://bambulab.com","contact":"support@bambulab.com"},
    {"no":1093,"category":"テクノロジー・ガジェット","name":"DJI Osmo Mobile 7P Phone Gimbal Stabilizer","maker":"DJI","ec_site":"DJI公式サイト","prod_url":"https://www.dji.com/osmo-mobile-7p","maker_hp":"https://www.dji.com","contact":"support@dji.com"},
    {"no":1094,"category":"テクノロジー・ガジェット","name":"Insta360 X5 360 Action Camera","maker":"Insta360","ec_site":"Insta360公式サイト","prod_url":"https://www.insta360.com/product/insta360-x5","maker_hp":"https://www.insta360.com","contact":"support@insta360.com"},
    {"no":1095,"category":"テクノロジー・ガジェット","name":"Elgato Stream Deck MK.2 Studio Controller","maker":"Elgato","ec_site":"Elgato公式サイト","prod_url":"https://www.elgato.com/en/stream-deck-mk2","maker_hp":"https://www.elgato.com","contact":"support@elgato.com"},
    # ── 美容続き ─────────────────────────────────────
    {"no":1096,"category":"美容・スキンケア","name":"LightStim for Wrinkles LED Light Therapy Device","maker":"LightStim","ec_site":"LightStim公式サイト","prod_url":"https://lightstim.com/products/lightstim-for-wrinkles","maker_hp":"https://lightstim.com","contact":"customercare@lightstim.com"},
    {"no":1097,"category":"美容・スキンケア","name":"Dermaflash LUXE+ Dermaplaning Device","maker":"Dermaflash","ec_site":"Dermaflash公式サイト","prod_url":"https://www.dermaflash.com/products/luxe-plus","maker_hp":"https://www.dermaflash.com","contact":"customerservice@dermaflash.com"},
    {"no":1098,"category":"美容・スキンケア","name":"Clarisonic Mia Smart Sonic Cleansing Device","maker":"Clarisonic","ec_site":"Sephora","prod_url":"https://www.sephora.com/brand/clarisonic","maker_hp":"https://www.clarisonic.com","contact":"support@clarisonic.com"},
    {"no":1099,"category":"美容・スキンケア","name":"Joanna Vargas Magic Glow Wand at-Home Facial","maker":"Joanna Vargas","ec_site":"Joanna Vargas公式サイト","prod_url":"https://joannavargas.com/collections/devices","maker_hp":"https://joannavargas.com","contact":"info@joannavargas.com"},
    {"no":1100,"category":"美容・スキンケア","name":"Braun Silk-expert Pro 5 IPL Hair Removal Device","maker":"Braun","ec_site":"Braun公式サイト","prod_url":"https://www.braun.com/en-us/ipl-hair-removal.html","maker_hp":"https://www.braun.com","contact":"support@braun.com"},
    # ── ファッション続き ───────────────────────────────
    {"no":1101,"category":"ファッション・アクセサリー","name":"Incase ICON Pack Backpack 22L MacBook Bag","maker":"Incase","ec_site":"Incase公式サイト","prod_url":"https://www.incase.com/collections/backpacks","maker_hp":"https://www.incase.com","contact":"support@incase.com"},
    {"no":1102,"category":"ファッション・アクセサリー","name":"Tropicfeel Nest Modular Travel Backpack 32L","maker":"Tropicfeel","ec_site":"Tropicfeel公式サイト","prod_url":"https://www.tropicfeel.com/collections/backpacks","maker_hp":"https://www.tropicfeel.com","contact":"hello@tropicfeel.com"},
    {"no":1103,"category":"ファッション・アクセサリー","name":"Knack Pack Series 2 Expandable Backpack","maker":"Knack Bags","ec_site":"Knack公式サイト","prod_url":"https://knackbags.com/collections/bags","maker_hp":"https://knackbags.com","contact":"info@knackbags.com"},
    {"no":1104,"category":"ファッション・アクセサリー","name":"Hyper Juice 245W USB-C Laptop Charger","maker":"Hyper","ec_site":"Hyper公式サイト","prod_url":"https://www.hypershop.com/collections/chargers","maker_hp":"https://www.hypershop.com","contact":"support@hypershop.com"},
    {"no":1105,"category":"クリーニング・収納・整理","name":"Woolite Dry Cleaner's Secret Dryer Activated Cleaning","maker":"Woolite","ec_site":"Target","prod_url":"https://www.target.com/s?searchTerm=woolite+dry+cleaner","maker_hp":"https://www.woolite.com","contact":"consumercare@rbnainfo.com"},
]

# ===================== Styles =====================
cat_fills = {
    "キッチン・調理器具":             PatternFill("solid", start_color="2E75B6"),
    "スマートホーム・インテリア・照明": PatternFill("solid", start_color="538135"),
    "ウェアラブル・ヘルス・フィットネス": PatternFill("solid", start_color="7030A0"),
    "アウトドア・スポーツ・旅行":       PatternFill("solid", start_color="C55A11"),
    "ペット用品":                      PatternFill("solid", start_color="843C0C"),
    "テクノロジー・ガジェット":         PatternFill("solid", start_color="244185"),
    "美容・スキンケア":                PatternFill("solid", start_color="C00000"),
    "子供・教育":                      PatternFill("solid", start_color="375623"),
    "ファッション・アクセサリー":       PatternFill("solid", start_color="7B3F00"),
    "クリーニング・収納・整理":         PatternFill("solid", start_color="404040"),
}
white_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
normal_font = Font(name="Arial", size=10)
link_font   = Font(name="Arial", size=10, color="0563C1", underline="single")
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill    = PatternFill("solid", start_color="EBF3FB")
center_va   = Alignment(vertical="center", wrap_text=True)

wb = load_workbook(r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸905件.xlsx")
ws = wb.active

for p in new_products:
    row_num = p["no"] + 1  # row 1 = header
    bg = alt_fill if row_num % 2 == 0 else None
    cat = p["category"]
    cat_fill = cat_fills.get(cat)

    contact_val = str(p["contact"]).strip()
    has_email = "@" in contact_val and not contact_val.startswith("http") and "問い合わせ" not in contact_val

    # Build row values
    values = [p["no"], cat, p["name"], p["maker"], p["ec_site"], p["prod_url"], p["maker_hp"], contact_val]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = border
        cell.alignment = center_va

        # Category column (col 2) → colored background + white bold
        if col_idx == 2:
            if cat_fill:
                cell.fill = cat_fill
                cell.font = white_font
            else:
                cell.font = normal_font
        # URL columns (col 6, 7) → hyperlink
        elif col_idx in (6, 7) and val and str(val).startswith("http"):
            cell.hyperlink = str(val)
            cell.font = link_font
            if bg and col_idx != 2:
                cell.fill = bg
        # Email / contact column (col 8)
        elif col_idx == 8:
            if has_email:
                gmail_url = build_gmail(contact_val, p["maker"], p["name"])
                cell.hyperlink = gmail_url
                cell.font = link_font
            elif "問い合わせ" in contact_val:
                url_part = contact_val.replace("（問い合わせページ）", "").strip()
                if url_part.startswith("http"):
                    cell.hyperlink = url_part
                    cell.value = "問い合わせページ"
                    cell.font = link_font
            if bg:
                cell.fill = bg
        else:
            cell.font = normal_font
            if bg and col_idx != 2:
                cell.fill = bg

output_path = r"C:\Users\hiro\HP0524\海外便利グッズリスト_日本未上陸1105件.xlsx"
wb.save(output_path)
print(f"完了: {output_path}")
print(f"追加件数: {len(new_products)} 件 (No.906〜1105)")
email_count = sum(1 for p in new_products if "@" in str(p["contact"]) and not str(p["contact"]).startswith("http"))
print(f"メールアドレスあり: {email_count} 件 ({email_count*100//len(new_products)}%)")
